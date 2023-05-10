import numpy as np
import openpyxl


def p(x):
    return -1
    # return 1
    # return -1


def f(x):
    return 1
    # return 2*x
    # return 2*x-np.pi


def exact(x):
    return -np.cos(x)-np.sin(x)+1 # 754
    # return np.sinh(x)/np.sinh(1) - 2*x # 751
    # return 2*x-np.pi+np.pi*np.cos(x)+2.2*np.sin(x) # 756


def getA(x, h):
    return 1/h**2-1/12*p(x-h)


def getB(x, h):
    return -2/h**2-5/6*p(x)


def getC(x, h):
    return 1/h**2-1/12*p(x+h)


def getD(x, h):
    return f(x)+1/12*(f(x+h)-1*f(x)+f(x-h))


def getAcc(x):
    y_exact = []
    for i in range(len(x)):
        y_exact.append(exact(x[i]))
    return y_exact


def progonka(a, b, f, p, h, y0, yn):
    k = a + h
    x = []
    while k < b - h:
        x.append(k)
        k += h

    A = np.zeros(len(x))
    B = np.zeros(len(x))
    C = np.zeros(len(x))
    D = np.zeros(len(x))

    A[0] = 0
    B[0] = getB(x[0], h)
    C[0] = getC(x[0], h)
    D[0] = getD(x[0], h) - y0*getA(x[0], h)
    for i in range(1, len(x)-1):
        A[i] = getA(x[i], h)
        B[i] = getB(x[i], h)
        C[i] = getC(x[i], h)
        D[i] = getD(x[i], h)
    A[len(x)-1] = getA(x[len(x)-1], h)
    B[len(x)-1] = getB(x[len(x)-1], h)
    C[len(x)-1] = 0
    D[len(x)-1] = getD(x[len(x)-1], h) - yn*getC(x[len(x)-1], h)
    omega = np.zeros(len(x))
    alpha = np.zeros(len(x))
    beta = np.zeros(len(x))

    omega[0] = B[0]
    alpha[0] = -C[0]/omega[0]
    beta[0] = D[0]/omega[0]

    for i in range(1, len(x) - 1):
        omega[i] = B[i] + A[i]*alpha[i-1]
        alpha[i] = -C[i]/omega[i]
        beta[i] = (D[i] - A[i]*beta[i-1])/omega[i]

    omega[len(x)-1] = B[len(x)-1] + A[len(x)-1]*alpha[len(x)-2]
    alpha[len(x)-1] = 0
    beta[len(x)-1] = (D[len(x)-1] - A[len(x)-1]*beta[len(x)-2])/omega[len(x)-1]

    y = np.zeros(len(x))
    y[len(y)-1] = beta[len(beta)-1]
    for i in range(len(y)-2, -1, -1):
        y[i] = alpha[i]*y[i+1] + beta[i]

    resultX = np.zeros(len(x)+2)
    resultY = np.zeros(len(x)+2)
    resultX[0] = a
    resultY[0] = y0
    for i in range(1, len(resultX)-1):
        resultX[i] = x[i-1]
        resultY[i] = y[i-1]
    resultX[len(resultX)-1] = b
    resultY[len(resultY)-1] = yn
    resultAcc = getAcc(resultX)
    Accuracy = abs(resultAcc - resultY)
    return resultX, resultY, resultAcc, Accuracy


a = 0
b = np.pi/2
y0 = 0
yn = 0
h = 0.01

# a = 0
# b = 3
# y0 = 0
# yn = exact(b)
# h = 0.05

# a = 0
# b = np.pi
# y0 = 0
# yn = 0
# h = 0.01


x, y_numerov, y_exact, acc = progonka(a, b, f, p, h, y0, yn)
x_h2, y_numerov_h2, y_exact_h2, acc_h2 = progonka(a, b, f, p, h/2, y0, yn)

workbook = openpyxl.Workbook()

worksheet = workbook.active
worksheet.cell(row=1, column=1, value='x')
worksheet.cell(row=1, column=2, value='y_numerov')
worksheet.cell(row=1, column=3, value='y_exact')
worksheet.cell(row=1, column=4, value='accuracy')
worksheet.cell(row=1, column=5, value='x_h2')
worksheet.cell(row=1, column=6, value='y_numerov_h2')
worksheet.cell(row=1, column=7, value='y_exact_h2')
worksheet.cell(row=1, column=8, value='accuracy_h2')

for i in range(len(x)):
    worksheet.cell(row=i+2, column=1, value=x[i])
    worksheet.cell(row=i+2, column=2, value=y_numerov[i])
    worksheet.cell(row=i+2, column=3, value=y_exact[i])
    worksheet.cell(row=i+2, column=4, value=acc[i])

for i in range(len(x_h2)):
    worksheet.cell(row=i+2, column=5, value=x_h2[i])
    worksheet.cell(row=i+2, column=6, value=y_numerov_h2[i])
    worksheet.cell(row=i+2, column=7, value=y_exact_h2[i])
    worksheet.cell(row=i+2, column=8, value=acc_h2[i])

workbook.save('output.xlsx')
